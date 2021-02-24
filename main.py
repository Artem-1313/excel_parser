from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
import os
import itertools

mypath = os.listdir(os.path.dirname(os.path.abspath(__file__)))
path = os.path.join(os.getcwd(), "terms_to_delete", "terms to delete lines.xlsx")
onlyfiles = [f for f in mypath if f.endswith(".xlsx") or f.endswith(".xls")]
print(onlyfiles)

def parse_excel_file(excel_file):
    def get_symbol(excel_file):
        workbook = load_workbook(filename=excel_file)
        sheet = workbook.active
        string = ''
        for row in sheet['A']:
            string += str(row.value)
        return string

    def check_word(word):
        spec_symbols = get_symbol(path)
        match = [l in spec_symbols for l in word]
        group = [k for k, g in itertools.groupby(match)]
        return sum(group) == 1

    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    max_row = sheet.max_row
    max_col = sheet.max_column
    i = 0
    # print(max_row)
    while i <= max_row:
        for row in range(max_col):
            i += 1
            for col in range(max_row):
                cellValue = str(sheet[get_column_letter(row + 1) + str(col + 1)].value)
                if check_word(cellValue) == 1:
                    #print(cellValue)
                     sheet.delete_rows(col + 1, 1)
    workbook.save(excel_file)

def brute_files(files):
    for i in files:
        parse_excel_file(i)

brute_files(onlyfiles)
#parse_excel_file("hello_world.xlsx")
